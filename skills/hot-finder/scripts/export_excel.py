#!/usr/bin/env python3
"""
hot-finder export_excel.py v2
将搜索结果导出为格式化 Excel (.xlsx)，支持多内容类型（视频/帖子/新闻）
"""
import json, sys, os, subprocess
from datetime import datetime

def ensure_openpyxl():
    try:
        import openpyxl
    except ImportError:
        print("⚠️  安装 openpyxl...", file=sys.stderr)
        subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "-q"], check=True)

def export_excel(results, output_path, keyword=""):
    ensure_openpyxl()
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "爆款内容"

    # 样式
    hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True, size=11)
    ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lft = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="DDDDDD")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 平台颜色
    colors = {
        "HackerNews": "FFF3E0", "Reddit": "FCE4EC", "Reddit/AI": "FCE4EC",
        "哔哩哔哩": "E3F2FD", "YouTube": "FFEBEE", "TikTok": "F3E5F5",
        "必应视频": "E8EAF6", "微博": "FFFDE7", "小红书": "FCE4EC",
    }

    headers = ["序号","平台","类型","标题","简介/摘要","评分/播放","评论数","作者","关键词","原文链接","讨论链接"]
    widths  = [5,    14,   8,   55,    40,          12,       8,    16,   10,   50,     40]

    for col,(h,w) in enumerate(zip(headers,widths),1):
        c = ws.cell(row=1,column=col,value=h)
        c.font=hdr_font; c.fill=hdr_fill; c.alignment=ctr; c.border=bdr
        ws.column_dimensions[get_column_letter(col)].width=w
    ws.row_dimensions[1].height=28
    ws.freeze_panes="A2"

    type_label = {"video":"🎬视频","post":"📝帖子","news":"📰新闻","tech_post":"💻技术"}

    for i,r in enumerate(results,1):
        row = i+1
        plat = r.get("platform","").split("/")[0].strip()
        fill = PatternFill(start_color=colors.get(plat,"FFFFFF"), end_color=colors.get(plat,"FFFFFF"), fill_type="solid")
        score = r.get("score") or r.get("plays","") or ""
        vals = [
            i,
            f"{r.get('icon','')} {r.get('platform','')}",
            type_label.get(r.get("content_type",""),"📄其他"),
            r.get("title",""),
            r.get("description","")[:300],
            str(score),
            str(r.get("comments","") or ""),
            r.get("author",""),
            r.get("keyword",""),
            r.get("link",""),
            r.get("hn_link","") or r.get("link",""),
        ]
        for col,val in enumerate(vals,1):
            c = ws.cell(row=row,column=col,value=val)
            c.fill=fill; c.border=bdr
            c.alignment=ctr if col in(1,3,6,7) else lft
            if col==4: c.font=Font(size=10,bold=True)
        ws.row_dimensions[row].height=45

    # 平台统计
    from collections import Counter
    ws2 = wb.create_sheet("来源统计")
    counter = Counter(r.get("platform","?") for r in results)
    type_counter = Counter(r.get("content_type","?") for r in results)
    ws2["A1"]="平台"; ws2["B1"]="条数"; ws2["C1"]="占比"
    for c in ["A1","B1","C1"]: ws2[c].font=Font(bold=True)
    total=len(results)
    for i,(plat,cnt) in enumerate(counter.most_common(),2):
        ws2.cell(i,1,plat); ws2.cell(i,2,cnt); ws2.cell(i,3,f"{cnt/total*100:.1f}%")
    ws2["E1"]="内容类型"; ws2["F1"]="条数"
    ws2["E1"].font=ws2["F1"].font=Font(bold=True)
    for i,(t,cnt) in enumerate(type_counter.most_common(),2):
        ws2.cell(i,5,type_label.get(t,t)); ws2.cell(i,6,cnt)
    for col in ["A","B","C","E","F"]: ws2.column_dimensions[col].width=18

    # 高分视频 sheet（score > 0 的 Top20）
    top_items = [r for r in results if r.get("score",0) > 0]
    top_items.sort(key=lambda x: -x.get("score",0))
    if top_items:
        ws3 = wb.create_sheet("热度排行")
        ws3.append(["排名","平台","标题","热度分","评论","链接"])
        for c in ws3[1]: c.font=Font(bold=True,color="FFFFFF"); c.fill=PatternFill(start_color="2E4057",end_color="2E4057",fill_type="solid")
        for i,r in enumerate(top_items[:30],1):
            ws3.append([i, f"{r.get('icon','')} {r.get('platform','')}", r.get("title",""), r.get("score",0), r.get("comments",""), r.get("link","")])
        ws3.column_dimensions["C"].width=60
        ws3.column_dimensions["F"].width=50

    # 元信息
    ws4 = wb.create_sheet("搜索信息")
    info = [("关键词",keyword),("生成时间",datetime.now().strftime("%Y-%m-%d %H:%M:%S")),("总条数",total),("平台数",len(counter)),("内容类型",", ".join(type_counter.keys()))]
    for row,( k,v) in enumerate(info,1):
        ws4.cell(row,1,k).font=Font(bold=True); ws4.cell(row,2,str(v))
    ws4.column_dimensions["A"].width=12; ws4.column_dimensions["B"].width=40

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    return output_path

if __name__=="__main__":
    import argparse
    p=argparse.ArgumentParser()
    p.add_argument("--input",required=True)
    p.add_argument("--output",required=True)
    p.add_argument("--keyword",default="")
    args=p.parse_args()
    with open(args.input,encoding="utf-8") as f: results=json.load(f)
    path=export_excel(results,args.output,args.keyword)
    print(f"✅ Excel 导出完成: {path} ({len(results)} 条, 4个Sheet)")
